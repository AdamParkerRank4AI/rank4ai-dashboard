#!/usr/bin/env python3
"""
ingest_gsc_drilldown.py — read GSC Coverage Drilldown XLSX exports from
~/Downloads/ and merge into src/data/live/gsc_coverage_drilldown.json.

GSC's Indexing → Pages → Why pages aren't indexed → click an issue → Export
gives a per-URL XLSX for that one issue. This script auto-detects which
property + issue from the file and the Metadata sheet, ingests the URL list,
and keeps a cumulative state across exports.

Filename pattern: <domain>-Coverage-Drilldown-YYYY-MM-DD*.xlsx

Output schema:
{
  "rank4ai": {
    "fetched_at": "...",
    "issues": {
      "Not found (404)": {
        "exported_at": "2026-05-08",
        "url_count": 25,
        "urls": [{"url": "...", "last_crawled": "..."}],
        "trend": [{"date": "...", "affected": N}, ...]
      },
      "Excluded by 'noindex' tag": {...},
      ...
    }
  }
}
"""
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

PROJECT_DIR = Path(__file__).resolve().parent.parent
LIVE = PROJECT_DIR / "src" / "data" / "live"
DOWNLOADS = Path.home() / "Downloads"
OUTPUT = LIVE / "gsc_coverage_drilldown.json"

# Map domain in filename → site_id. GSC names URL-prefix exports
# "<domain>-Coverage-Drilldown-..." and domain-property exports
# "https___<domain>_-Coverage-Drilldown-..." — so match the domain anywhere.
DOMAIN_TO_SITE = {
    "rank4ai.co.uk": "rank4ai",
    "marketinvoice.co.uk": "market-invoice",
    "seocompare.co.uk": "seocompare",
    "merchanthq.co.uk": "cardmachines",
    "peptideclear.co.uk": "peptideclear",
    "fundbiz.co.uk": "fundbiz",
    "bestbusinessloans.ai": "bestbusinessloans",
    "kartapay.com": "kartapay",
}


def detect_site(filename):
    name = os.path.basename(filename).lower()
    # Longest domain first so e.g. a sub-brand never shadows the full domain.
    for domain in sorted(DOMAIN_TO_SITE, key=len, reverse=True):
        if domain in name:
            return DOMAIN_TO_SITE[domain], domain
    return None, None


def parse_xlsx(path):
    """Returns (issue_label, url_rows, trend_rows)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    issue = None
    if "Metadata" in wb.sheetnames:
        ws = wb["Metadata"]
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == "Issue":
                issue = row[1]
                break
    urls = []
    if "Table" in wb.sheetnames:
        ws = wb["Table"]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if row and row[0]:
                last = row[1]
                last_str = last.strftime("%Y-%m-%d") if hasattr(last, "strftime") else str(last) if last else None
                urls.append({"url": row[0], "last_crawled": last_str})
    trend = []
    if "Chart" in wb.sheetnames:
        ws = wb["Chart"]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if row and row[0]:
                d = row[0]
                d_str = d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d)
                trend.append({"date": d_str, "affected": int(row[1] or 0) if row[1] is not None else 0})
    return issue, urls, trend


def load_existing():
    if OUTPUT.exists():
        try:
            with open(OUTPUT) as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def main():
    # Find all GSC Coverage XLSX files in Downloads
    candidates = []
    for f in DOWNLOADS.glob("*-Coverage-Drilldown-*.xlsx"):
        candidates.append(f)
    if len(sys.argv) > 1:
        # Allow explicit paths to be passed
        candidates = [Path(p) for p in sys.argv[1:] if p.endswith(".xlsx")]

    if not candidates:
        print("No GSC Coverage Drilldown XLSX files found in ~/Downloads")
        return

    state = load_existing()
    ingested = 0
    for path in sorted(candidates):
        site_id, domain = detect_site(path.name)
        if not site_id:
            print(f"  skip {path.name} (unknown domain)")
            continue
        issue, urls, trend = parse_xlsx(path)
        if not issue:
            print(f"  skip {path.name} (no Issue in metadata)")
            continue
        state.setdefault(site_id, {"issues": {}})
        state[site_id]["fetched_at"] = datetime.now(timezone.utc).isoformat()
        state[site_id]["domain"] = domain
        # Newest export wins per (site, issue)
        prior = state[site_id]["issues"].get(issue, {})
        prior_date = prior.get("exported_at", "")
        # Try to extract export date from filename
        m = re.search(r"(\d{4}-\d{2}-\d{2})", path.name)
        export_date = m.group(1) if m else datetime.now().strftime("%Y-%m-%d")
        if export_date >= prior_date:
            state[site_id]["issues"][issue] = {
                "exported_at": export_date,
                "url_count": len(urls),
                "urls": urls,
                "trend": trend,
                "source_file": path.name,
            }
            ingested += 1
            print(f"  {site_id} · {issue}: {len(urls)} URLs (from {path.name})")
        else:
            print(f"  skip {path.name} — older than existing {issue} ({prior_date})")

    OUTPUT.write_text(json.dumps(state, indent=2))
    print(f"\nIngested {ingested} drilldowns → {OUTPUT}")
    # Per-site summary
    for site_id, data in state.items():
        issues = data.get("issues") or {}
        total = sum(i["url_count"] for i in issues.values())
        print(f"  {site_id}: {len(issues)} issue types, {total} URLs total")
        for label, d in sorted(issues.items(), key=lambda x: -x[1]["url_count"]):
            print(f"    {d['url_count']:>4}  {label}  (exported {d['exported_at']})")


if __name__ == "__main__":
    main()
